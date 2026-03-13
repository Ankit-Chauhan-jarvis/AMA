import os
import re
import json
import time
import argparse
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from collections import Counter, defaultdict

from groq import Groq
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_POS_BG      = "C6EFCE"   # green
C_POS_FG      = "276221"
C_NEG_BG      = "FFC7CE"   # red
C_NEG_FG      = "9C0006"
C_NEU_BG      = "FFEB9C"   # amber
C_NEU_FG      = "9C6500"
C_ALT_ROW     = "EEF2FF"   # light blue-grey for alternating rows
C_SECTION_BG  = "D9E1F2"   # section sub-header

SENTIMENT_LABELS = ["positive", "neutral", "negative"]

SYSTEM_PROMPT = (
    "You are an expert meeting analyst performing sentiment analysis on meeting transcript utterances.\n"
    "For each utterance, assess the speaker's emotional tone and communicative sentiment.\n\n"
    "Guidelines:\n"
    "  - positive : enthusiasm, agreement, encouragement, progress, satisfaction\n"
    "  - neutral  : factual statements, questions, procedural talk, balanced discussion\n"
    "  - negative : frustration, disagreement, confusion, concern, hesitation, criticism\n\n"
    "Return ONLY a valid JSON array. No markdown, no extra text.\n"
    "Each element: {\"id\": <int>, \"sentiment\": \"<label>\", \"score\": <float -1.0 to 1.0>, \"reasoning\": \"<1 sentence>\"}\n"
    "  score: +1.0 = very positive, 0.0 = neutral, -1.0 = very negative\n"
    "  reasoning: concise explanation of why you chose that sentiment.\n"
    "You MUST return an entry for EVERY id provided."
)


def extract_json_array(text: str) -> List[Dict[str, Any]]:
    start = text.find("[")
    end = text.rfind("]")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("Model output did not contain a JSON array.")
    return json.loads(text[start:end + 1])


def _call_llm_sentiment_batch(
    client: Groq,
    model: str,
    batch_items: List[Tuple[int, str]],
    temperature: float,
) -> Dict[int, Dict[str, Any]]:
    payload = [
        {"id": sid, "text": re.sub(r"\s+", " ", (text or "")).strip()}
        for sid, text in batch_items
    ]
    user_prompt = (
        "Analyse the sentiment of each meeting utterance.\n"
        "Return ONLY a JSON array — no markdown, no extra text.\n"
        "Each element: {\"id\": <int>, \"sentiment\": \"positive|neutral|negative\", "
        "\"score\": <float -1.0 to 1.0>, \"reasoning\": \"<1 sentence>\"}\n\n"
        "INPUT:\n" + json.dumps(payload, ensure_ascii=False)
    )
    completion = client.chat.completions.create(
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        model=model,
        temperature=temperature,
        max_tokens=2048,
    )
    raw = (completion.choices[0].message.content or "").strip()
    arr = extract_json_array(raw)
    result: Dict[int, Dict[str, Any]] = {}
    for obj in arr:
        sid = int(obj["id"])
        sentiment = str(obj.get("sentiment", "")).strip().lower()
        if sentiment not in SENTIMENT_LABELS:
            raise ValueError(f"Invalid sentiment '{sentiment}' for id={sid}")
        score = max(-1.0, min(1.0, float(obj.get("score", 0.0))))
        result[sid] = {
            "sentiment": sentiment,
            "score": round(score, 3),
            "reasoning": str(obj.get("reasoning", "")).strip(),
        }
    return result


def analyse_batch(
    client: Groq,
    model: str,
    batch_items: List[Tuple[int, str]],
    temperature: float = 0.0,
    max_retries: int = 3,
) -> Dict[int, Dict[str, Any]]:
    expected_ids = [sid for sid, _ in batch_items]
    expected_set = set(expected_ids)
    collected: Dict[int, Dict[str, Any]] = {}
    remaining = list(batch_items)

    for attempt in range(1, max_retries + 1):
        try:
            got = _call_llm_sentiment_batch(client, model, remaining, temperature)
            for sid, data in got.items():
                if sid in expected_set:
                    collected[sid] = data
            missing = [sid for sid in expected_ids if sid not in collected]
            if not missing:
                break
            remaining = [(sid, text) for sid, text in batch_items if sid in set(missing)]
            time.sleep(0.75 * attempt)
        except Exception as e:
            print(f"  [WARN] Batch attempt {attempt} failed: {e}")
            time.sleep(1.0 * attempt)

    for sid, text in remaining:
        if sid in collected:
            continue
        for attempt in range(1, 5):
            try:
                got = _call_llm_sentiment_batch(client, model, [(sid, text)], temperature)
                if sid in got:
                    collected[sid] = got[sid]
                    break
            except Exception:
                time.sleep(1.5 * attempt)
        if sid not in collected:
            print(f"  [ERROR] Could not get sentiment for id={sid}. Defaulting to neutral.")
            collected[sid] = {"sentiment": "neutral", "score": 0.0, "reasoning": "Could not analyse."}

    return collected


# ── Transcript loading ────────────────────────────────────────────────────────

def load_transcript(json_path: str) -> List[Dict[str, Any]]:
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("Expected JSON to be a list of utterances.")
    utterances = []
    for i, item in enumerate(data, start=1):
        text = str(item.get("text", "")).strip()
        if not text:
            continue
        utterances.append({
            "id": i,
            "speaker": str(item.get("speaker", "UNKNOWN")),
            "start": float(item.get("start", 0.0)),
            "end": float(item.get("end", 0.0)),
            "text": text,
        })
    return utterances


# ── Aggregations ──────────────────────────────────────────────────────────────

def aggregate_speaker_level(utterances: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    speaker_scores: Dict[str, List[float]] = defaultdict(list)
    speaker_sentiments: Dict[str, List[str]] = defaultdict(list)
    for u in utterances:
        spk = u["speaker"]
        speaker_scores[spk].append(u["score"])
        speaker_sentiments[spk].append(u["sentiment"])
    result = {}
    for spk in speaker_scores:
        scores = speaker_scores[spk]
        sentiments = speaker_sentiments[spk]
        mean_score = round(sum(scores) / len(scores), 3)
        dominant = Counter(sentiments).most_common(1)[0][0]
        dist = dict(Counter(sentiments))
        result[spk] = {
            "utterance_count": len(scores),
            "mean_score": mean_score,
            "dominant_sentiment": dominant,
            "positive": dist.get("positive", 0),
            "neutral": dist.get("neutral", 0),
            "negative": dist.get("negative", 0),
        }
    return result


def aggregate_temporal_arc(utterances: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    n = len(utterances)
    if n == 0:
        return {}
    thirds = {
        "Opening": utterances[: n // 3],
        "Middle":  utterances[n // 3: 2 * n // 3],
        "Closing": utterances[2 * n // 3:],
    }
    arc = {}
    for phase, items in thirds.items():
        if not items:
            continue
        scores = [u["score"] for u in items]
        sentiments = [u["sentiment"] for u in items]
        dist = dict(Counter(sentiments))
        arc[phase] = {
            "utterance_count": len(items),
            "mean_score": round(sum(scores) / len(scores), 3),
            "dominant_sentiment": Counter(sentiments).most_common(1)[0][0],
            "positive": dist.get("positive", 0),
            "neutral": dist.get("neutral", 0),
            "negative": dist.get("negative", 0),
            "time_range": f"{items[0]['start']:.1f}s – {items[-1]['end']:.1f}s",
        }
    return arc


def overall_stats(utterances: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not utterances:
        return {}
    scores = [u["score"] for u in utterances]
    sentiments = [u["sentiment"] for u in utterances]
    dist = dict(Counter(sentiments))
    return {
        "total_utterances": len(utterances),
        "mean_score": round(sum(scores) / len(scores), 3),
        "min_score": round(min(scores), 3),
        "max_score": round(max(scores), 3),
        "dominant_sentiment": Counter(sentiments).most_common(1)[0][0],
        "positive": dist.get("positive", 0),
        "neutral": dist.get("neutral", 0),
        "negative": dist.get("negative", 0),
    }


def generate_holistic_summary(
    client: Groq,
    model: str,
    utterances: List[Dict[str, Any]],
    speaker_stats: Dict[str, Any],
    temporal_arc: Dict[str, Any],
    stats: Dict[str, Any],
) -> str:
    top_positive = sorted(utterances, key=lambda u: u["score"], reverse=True)[:3]
    top_negative = sorted(utterances, key=lambda u: u["score"])[:3]
    context = {
        "overall_stats": stats,
        "temporal_arc": temporal_arc,
        "speaker_level": speaker_stats,
        "most_positive_moments": [
            {"speaker": u["speaker"], "text": u["text"][:120], "score": u["score"]} for u in top_positive
        ],
        "most_negative_moments": [
            {"speaker": u["speaker"], "text": u["text"][:120], "score": u["score"]} for u in top_negative
        ],
    }
    prompt = (
        "You are an expert meeting analyst. Based on the quantitative sentiment analysis results below, "
        "write a concise (3–5 paragraph) meeting sentiment report covering:\n"
        "  1. Overall meeting tone and sentiment\n"
        "  2. How sentiment evolved across the meeting (opening / middle / closing)\n"
        "  3. Notable speaker-level sentiment patterns\n"
        "  4. Key positive and negative moments\n"
        "  5. Actionable interpretation for the team\n\n"
        "Write in clear, professional prose. Do not repeat raw numbers excessively.\n\n"
        "ANALYSIS DATA:\n" + json.dumps(context, indent=2, ensure_ascii=False)
    )
    completion = client.chat.completions.create(
        messages=[{"role": "user", "content": prompt}],
        model=model,
        temperature=0.3,
        max_tokens=1024,
    )
    return (completion.choices[0].message.content or "").strip()


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _header_style(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, size=11):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _section_style(cell, text=""):
    cell.value = text
    cell.font = Font(name="Arial", bold=True, color="1F3864", size=10)
    cell.fill = PatternFill("solid", fgColor=C_SECTION_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _sentiment_fill(sentiment: str) -> PatternFill:
    mapping = {
        "positive": (C_POS_BG, C_POS_FG),
        "negative": (C_NEG_BG, C_NEG_FG),
        "neutral":  (C_NEU_BG, C_NEU_FG),
    }
    bg, fg = mapping.get(sentiment, ("FFFFFF", "000000"))
    return PatternFill("solid", fgColor=bg)

def _sentiment_font(sentiment: str) -> Font:
    mapping = {
        "positive": C_POS_FG,
        "negative": C_NEG_FG,
        "neutral":  C_NEU_FG,
    }
    return Font(name="Arial", color=mapping.get(sentiment, "000000"), bold=True, size=10)

def _score_bar(score: float) -> str:
    """Convert score (-1 to 1) to a simple text bar."""
    blocks = int(round((score + 1) / 2 * 10))
    return "█" * blocks + "░" * (10 - blocks)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _freeze_and_filter(ws, cell="A2"):
    ws.freeze_panes = cell
    ws.auto_filter.ref = ws.dimensions


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary_sheet(ws, stats, summary_text, meeting_name):
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Meeting Sentiment Analysis — {meeting_name}"
    title_cell.font = Font(name="Arial", bold=True, size=16, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.row_dimensions[1].height = 36

    ws.append([])  # blank row

    # ── Overall stats block ──
    _section_style(ws["A3"], "OVERALL MEETING STATS")
    ws.merge_cells("A3:F3")
    ws.row_dimensions[3].height = 22

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        _header_style(c)
    ws.row_dimensions[4].height = 20

    stat_rows = [
        ("Total Utterances",  stats["total_utterances"]),
        ("Overall Sentiment", stats["dominant_sentiment"].title()),
        ("Mean Score",        stats["mean_score"]),
        ("Min Score",         stats["min_score"]),
        ("Max Score",         stats["max_score"]),
        ("Positive Utterances", stats["positive"]),
        ("Neutral Utterances",  stats["neutral"]),
        ("Negative Utterances", stats["negative"]),
    ]
    for r, (label, val) in enumerate(stat_rows, 5):
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = Font(name="Arial", size=10)
        if label == "Overall Sentiment":
            cell.fill = _sentiment_fill(val.lower())
            cell.font = _sentiment_font(val.lower())
        if r % 2 == 0:
            ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=C_ALT_ROW)
        for col in [1, 2]:
            ws.cell(row=r, column=col).border = _thin_border()
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="left", vertical="center")

    # ── Holistic summary block ──
    summary_start = 5 + len(stat_rows) + 2
    ws.cell(row=summary_start, column=1).value = "HOLISTIC SUMMARY"
    _section_style(ws.cell(row=summary_start, column=1), "HOLISTIC SUMMARY")
    ws.merge_cells(f"A{summary_start}:F{summary_start}")
    ws.row_dimensions[summary_start].height = 22

    text_row = summary_start + 1
    ws.merge_cells(f"A{text_row}:F{text_row + 25}")
    tc = ws.cell(row=text_row, column=1, value=summary_text)
    tc.font = Font(name="Arial", size=10)
    tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[text_row].height = 300

    for col, width in zip("ABCDEF", [28, 18, 18, 18, 18, 18]):
        _set_col_width(ws, col, width)


def build_utterances_sheet(ws, utterances):
    ws.sheet_view.showGridLines = False

    headers = ["#", "Speaker", "Start (s)", "End (s)", "Sentiment", "Score", "Score Bar", "Reasoning", "Text"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        _header_style(c)
    ws.row_dimensions[1].height = 22

    for row_idx, u in enumerate(utterances, 2):
        sentiment = u["sentiment"]
        score = u["score"]
        alt = (row_idx % 2 == 0)

        values = [
            u["id"], u["speaker"], round(u["start"], 1), round(u["end"], 1),
            sentiment.title(), score, _score_bar(score), u["reasoning"], u["text"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in [8, 9]))
            cell.border = _thin_border()
            if alt:
                cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)

        # Colour-code sentiment cell
        sent_cell = ws.cell(row=row_idx, column=5)
        sent_cell.fill = _sentiment_fill(sentiment)
        sent_cell.font = _sentiment_font(sentiment)
        sent_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Score cell — right-align
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="top")

        # Score bar — monospace
        ws.cell(row=row_idx, column=7).font = Font(name="Courier New", size=9)

        ws.row_dimensions[row_idx].height = 45

    col_widths = [5, 14, 10, 10, 12, 8, 14, 40, 70]
    for col, width in enumerate(col_widths, 1):
        _set_col_width(ws, get_column_letter(col), width)

    _freeze_and_filter(ws, "A2")


def build_speakers_sheet(ws, speaker_stats):
    ws.sheet_view.showGridLines = False

    headers = ["Speaker", "Utterances", "Mean Score", "Score Bar", "Dominant Sentiment",
               "Positive", "Neutral", "Negative"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        _header_style(c)
    ws.row_dimensions[1].height = 22

    # Sort by mean score descending
    sorted_spk = sorted(speaker_stats.items(), key=lambda x: x[1]["mean_score"], reverse=True)

    for row_idx, (spk, data) in enumerate(sorted_spk, 2):
        dominant = data["dominant_sentiment"]
        values = [
            spk, data["utterance_count"], data["mean_score"],
            _score_bar(data["mean_score"]),
            dominant.title(),
            data["positive"], data["neutral"], data["negative"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)

        # Score bar monospace
        ws.cell(row=row_idx, column=4).font = Font(name="Courier New", size=10)

        # Dominant sentiment coloured
        sent_cell = ws.cell(row=row_idx, column=5)
        sent_cell.fill = _sentiment_fill(dominant)
        sent_cell.font = _sentiment_font(dominant)

        # Colour positive/neutral/negative counts
        for col_idx, sentiment in [(6, "positive"), (7, "neutral"), (8, "negative")]:
            c = ws.cell(row=row_idx, column=col_idx)
            if c.value and c.value > 0:
                c.fill = _sentiment_fill(sentiment)
                c.font = _sentiment_font(sentiment)

        ws.row_dimensions[row_idx].height = 22

    # Excel SUM formulas at bottom
    last_data_row = 1 + len(sorted_spk)
    summary_row = last_data_row + 2
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=summary_row, column=2).value = f"=SUM(B2:B{last_data_row})"
    ws.cell(row=summary_row, column=6).value = f"=SUM(F2:F{last_data_row})"
    ws.cell(row=summary_row, column=7).value = f"=SUM(G2:G{last_data_row})"
    ws.cell(row=summary_row, column=8).value = f"=SUM(H2:H{last_data_row})"
    for col in [1, 2, 6, 7, 8]:
        ws.cell(row=summary_row, column=col).fill = PatternFill("solid", fgColor=C_SECTION_BG)
        ws.cell(row=summary_row, column=col).font = Font(name="Arial", bold=True, size=10)

    col_widths = [16, 13, 12, 16, 20, 10, 10, 10]
    for col, width in enumerate(col_widths, 1):
        _set_col_width(ws, get_column_letter(col), width)

    ws.freeze_panes = "A2"


def build_arc_sheet(ws, temporal_arc):
    ws.sheet_view.showGridLines = False

    headers = ["Phase", "Time Range", "Utterances", "Mean Score", "Score Bar",
               "Dominant Sentiment", "Positive", "Neutral", "Negative"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        _header_style(c)
    ws.row_dimensions[1].height = 22

    phase_order = ["Opening", "Middle", "Closing"]
    for row_idx, phase in enumerate(phase_order, 2):
        if phase not in temporal_arc:
            continue
        data = temporal_arc[phase]
        dominant = data["dominant_sentiment"]
        values = [
            phase, data["time_range"], data["utterance_count"], data["mean_score"],
            _score_bar(data["mean_score"]),
            dominant.title(),
            data["positive"], data["neutral"], data["negative"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)

        ws.cell(row=row_idx, column=5).font = Font(name="Courier New", size=10)

        sent_cell = ws.cell(row=row_idx, column=6)
        sent_cell.fill = _sentiment_fill(dominant)
        sent_cell.font = _sentiment_font(dominant)

        for col_idx, sentiment in [(7, "positive"), (8, "neutral"), (9, "negative")]:
            c = ws.cell(row=row_idx, column=col_idx)
            if c.value and c.value > 0:
                c.fill = _sentiment_fill(sentiment)
                c.font = _sentiment_font(sentiment)

        ws.row_dimensions[row_idx].height = 22

    col_widths = [12, 22, 13, 12, 16, 20, 10, 10, 10]
    for col, width in enumerate(col_widths, 1):
        _set_col_width(ws, get_column_letter(col), width)

    ws.freeze_panes = "A2"


# ── Save Excel ────────────────────────────────────────────────────────────────

def save_excel(
    out_path: str,
    utterances: List[Dict[str, Any]],
    stats: Dict[str, Any],
    speaker_stats: Dict[str, Any],
    temporal_arc: Dict[str, Any],
    summary_text: str,
    meeting_name: str,
) -> None:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    build_summary_sheet(ws_summary, stats, summary_text, meeting_name)

    ws_utt = wb.create_sheet("Utterances")
    build_utterances_sheet(ws_utt, utterances)

    ws_spk = wb.create_sheet("Speakers")
    build_speakers_sheet(ws_spk, speaker_stats)

    ws_arc = wb.create_sheet("Temporal Arc")
    build_arc_sheet(ws_arc, temporal_arc)

    wb.save(out_path)
    print(f"Saved Excel: {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Meeting sentiment analysis via Groq LLM.")
    parser.add_argument("--transcript", required=True, help="Path to meeting JSON transcript.")
    parser.add_argument("--out", default="sentiment_output.xlsx", help="Output Excel path (.xlsx).")
    parser.add_argument("--batch_size", type=int, default=15, help="Utterances per LLM call.")
    parser.add_argument("--model", default=None, help="Groq model name (overrides GROQ_MODEL env).")
    args = parser.parse_args()

    load_dotenv()

    GROQ_API_KEY = os.getenv("GROQ_API_KEY")
    if not GROQ_API_KEY:
        raise ValueError("GROQ_API_KEY not set. Add it to your .env file.")

    model = args.model or os.getenv("GROQ_MODEL", "openai/gpt-oss-20b")
    client = Groq(api_key=GROQ_API_KEY)
    meeting_name = Path(args.transcript).stem

    # 1. Load
    print(f"Loading transcript: {args.transcript}")
    utterances = load_transcript(args.transcript)
    print(f"  {len(utterances)} utterances loaded.")

    # 2. Utterance-level sentiment
    print(f"\nRunning utterance-level sentiment (batch_size={args.batch_size})...")
    indexed: List[Tuple[int, str]] = [(u["id"], u["text"]) for u in utterances]
    id_to_sentiment: Dict[int, Dict[str, Any]] = {}
    total_batches = (len(indexed) + args.batch_size - 1) // args.batch_size

    for i in range(0, len(indexed), args.batch_size):
        batch = indexed[i: i + args.batch_size]
        batch_num = i // args.batch_size + 1
        print(f"  Batch {batch_num}/{total_batches} — ids {batch[0][0]}–{batch[-1][0]}")
        result = analyse_batch(client, model, batch, temperature=0.0)
        id_to_sentiment.update(result)

    for u in utterances:
        s = id_to_sentiment.get(u["id"], {"sentiment": "neutral", "score": 0.0, "reasoning": ""})
        u["sentiment"] = s["sentiment"]
        u["score"] = s["score"]
        u["reasoning"] = s["reasoning"]

    # 3. Aggregations
    print("\nAggregating results...")
    stats = overall_stats(utterances)
    speaker_stats = aggregate_speaker_level(utterances)
    temporal_arc = aggregate_temporal_arc(utterances)

    print(f"  Dominant sentiment : {stats['dominant_sentiment']}  (mean: {stats['mean_score']})")
    print(f"  Distribution       : +{stats['positive']}  ~{stats['neutral']}  -{stats['negative']}")

    # 4. Holistic summary
    print("\nGenerating holistic summary...")
    summary = generate_holistic_summary(client, model, utterances, speaker_stats, temporal_arc, stats)
    print("\n" + "=" * 60)
    print(summary)
    print("=" * 60)

    # 5. Save Excel
    out_path = args.out
    if not out_path.endswith(".xlsx"):
        out_path += ".xlsx"

    save_excel(
        out_path=out_path,
        utterances=utterances,
        stats=stats,
        speaker_stats=speaker_stats,
        temporal_arc=temporal_arc,
        summary_text=summary,
        meeting_name=meeting_name,
    )
    print(f"\nDone — {len(utterances)} utterances written to {out_path}")


if __name__ == "__main__":
    main()